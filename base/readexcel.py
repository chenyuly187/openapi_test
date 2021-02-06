# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from base.operation_excel import OperationExcel

class ReadExcel(object):
    def __init__(self,file):
        self.file = file
        self.OperaExcel = OperationExcel(self.file)

    def readTotalData(self,totalData=None):
        '''获取整个excel中sheet页面中的数据'''
        if totalData is None:
            totalData = []

        data = load_workbook(self.file)
        # 获取sheets 列表的名称
        sheet_names = data.sheetnames
        # 遍历sheet列表
        for sheet_name in sheet_names:
            # 获取最大行
            max_lines = data[sheet_name].max_row
            # 获取最大列
            max_column = data[sheet_name].max_column
            # row 下标从第二行开始，因为第一行是标题
            for row_index in range(2,max_lines+1):
                signle_line_value = []
                for column_index in range(1,max_column+1):
                    signle_line_value.append(data[sheet_name].cell(row=row_index,column=column_index).value)
                totalData.append(signle_line_value)

        return totalData
