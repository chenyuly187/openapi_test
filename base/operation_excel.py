# -*- coding: utf-8 -*-
from openpyxl import load_workbook

class OperationExcel(object):
    def __init__(self,file):
        self.file = file

    def get_total_data(self):
        '''读取整个文件'''
        data = load_workbook(self.file)

        return data

    def get_sheets_name(self):
        '''统计excel中有多少个sheet页面'''
        table = self.get_total_data()
        sheet_name = table.sheetnames

        return sheet_name

    def get_sheet_data(self,sheet_name):
        '''获取sheet页面中的数据'''
        data = self.get_total_data()

        return data[sheet_name]

    def get_excel_line(self,sheet_name):
        '''获取excel文件中的最大行数'''
        line = self.get_sheet_data(sheet_name)

        return line.max_row

    def get_excel_column(self,sheet_name):
        '''获取excel中的最大列'''
        column = self.get_sheet_data(sheet_name)

        return column.max_column























