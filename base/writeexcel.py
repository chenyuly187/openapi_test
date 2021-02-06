# -*- coding: utf-8 -*-
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Color,PatternFill

from config.excel_title import sheetName,ExcelTitle

class WriteToExcel(object):
    def __init__(self,data,file):
        self.data = data
        self.file = file
        self.workbook = Workbook()

        if os.path.exists(self.file):
            os.remove(self.file)

    def write_title_to_excel(self):
        '''写入excel标题'''
        for index,title in enumerate(ExcelTitle):
            sheet = self.workbook.create_sheet(sheetName[index])
            for col in range(1,len(title)+1):
                sheet.cell(row=1,column=col,value=title[col-1])
        self.workbook.save(self.file)

    def get_same_module_data(self,datas,module):
        '''获取同一模块的对话'''
        resp = []
        for data in datas:
            if data[0] == module:
                resp.append(data)
        return resp

    def write_result_to_excel(self):
        '''将测试结果写入到excel 中'''
        self.write_title_to_excel()
        rb = load_workbook(self.file)
        # 将sheet 空白页删除
        black_sheet = rb["Sheet"]
        rb.remove(black_sheet)
        for key in sheetName.keys():
            same_module_data = self.get_same_module_data(self.data,key)
            row = 1
            sheet = rb[sheetName[key]]
            for line in same_module_data:
                '''写入每行数据'''
                for col_index,row_value in enumerate(line):
                    # Fail 的单元格显示红色
                    if row_value == "Fail":
                        fill = PatternFill(patternType='solod',fill_type='solid', fgColor=Color('FF0000'))
                        # row 下标是从1开始的，但是在写入的时候，row 下标从2开始，因为第一行是标题
                        sheet.cell(row=row+1,column=col_index+1,value=row_value).fill = fill
                    else:
                        sheet.cell(row=row+1,column=col_index+1,value=row_value)
                row += 1
        rb.save(self.file)


if __name__ == "__main__":
    data = [[0,'Question1','SQ1','Answer1','Result1','Response_Answer1'],[0,'Question2','SQ2','Answer2','Result2','Response_Answer2'],[1,'Question3','SQ3','Answer3','Result3','Response_Answer3'],[1,'Question4','SQ4','Answer4','Result4','Response_Answer4']]
    file = "LY.xlsx"
    write_excel = WriteToExcel(data,file)
    write_excel.write_result_to_excel()